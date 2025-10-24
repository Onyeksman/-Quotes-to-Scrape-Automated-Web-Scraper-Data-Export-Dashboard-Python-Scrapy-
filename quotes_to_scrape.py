
import scrapy
from scrapy.crawler import CrawlerProcess
import re
import unicodedata
import html
import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class QuotesFinalSpider(scrapy.Spider):
    name = "quotes_final_clean"
    start_urls = ["http://quotes.toscrape.com/"]

    custom_settings = {
        "LOG_LEVEL": "INFO",
        "CONCURRENT_REQUESTS": 8,
        "AUTOTHROTTLE_ENABLED": True,
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.seq = 0
        self.items = {}
        self.csv_file = "quotestoscrape.csv"
        self.xlsx_file = "quotestoscrape.xlsx"

    # ---------- Clean Text ----------
    def clean_text(self, text):
        """Normalize, remove junk characters, replace newlines/tabs with space."""
        if not text:
            return ""
        text = html.unescape(text)
        text = unicodedata.normalize("NFKC", text)
        text = text.encode("ascii", "ignore").decode("ascii")
        text = re.sub(r"[\r\n\t]+", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def format_dob(self, dob_text):
        """Format DOB into YYYY-MM-DD."""
        dob_text = self.clean_text(dob_text)
        try:
            dt = datetime.strptime(dob_text, "%B %d, %Y")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return dob_text or ""

    # ---------- Parse Quotes ----------
    def parse(self, response):
        self.logger.info(f"Scraping page: {response.url}")

        for quote_sel in response.css("div.quote"):
            self.seq += 1
            seq = self.seq

            quote_text = self.clean_text(quote_sel.css('span[itemprop="text"]::text').get(default=""))
            author = self.clean_text(quote_sel.css('small[itemprop="author"]::text').get(default=""))
            tags_list = [self.clean_text(t) for t in quote_sel.css("div.tags a.tag::text").getall()]
            tags = ", ".join(tags_list) if tags_list else ""
            about_link = quote_sel.css("span a::attr(href)").get()
            about_url = response.urljoin(about_link) if about_link else ""

            if about_url:
                yield response.follow(
                    about_url,
                    callback=self.parse_author,
                    cb_kwargs={
                        "seq": seq,
                        "Quote": quote_text,
                        "Author": author,
                        "Tags": tags,
                        "About": about_url,
                    },
                    dont_filter=True,
                )
            else:
                self.items[seq] = {
                    "Author": author,
                    "Quote": quote_text,
                    "About the Author": "",
                    "DOB": "",
                    "Place of Birth": "",
                    "Bio": "",
                    "Tags": tags,
                }

        # ---------- Pagination: strictly sequential ----------
        next_page = response.css("li.next a::attr(href)").get()
        if next_page:
            yield scrapy.Request(
                url=response.urljoin(next_page),
                callback=self.parse,
                priority=-1,  # ensures sequential page order
                dont_filter=True,
            )

    # ---------- Parse Author ----------
    def parse_author(self, response, seq, Quote, Author, Tags, About):
        dob = self.format_dob(response.css("span.author-born-date::text").get(default=""))
        pob = self.clean_text(response.css("span.author-born-location::text").get(default="").replace("in ", ""))
        bio = self.clean_text(response.css("div.author-description::text").get(default=""))

        self.items[seq] = {
            "Author": Author,
            "Quote": Quote,
            "About the Author": About,
            "DOB": dob,
            "Place of Birth": pob,
            "Bio": bio,
            "Tags": Tags,
        }

    # ---------- Write CSV + XLSX ----------
    def closed(self, reason):
        keys = ["Author", "Quote", "About the Author", "DOB", "Place of Birth", "Bio", "Tags"]

        # Write CSV first
        with open(self.csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=keys, quoting=csv.QUOTE_ALL, quotechar='"')
            writer.writeheader()
            for seq in sorted(self.items.keys()):
                row = {k: self.items[seq].get(k, "") for k in keys}
                writer.writerow(row)

        # Read CSV and clean + prepare data
        rows = []
        seen = set()
        with open(self.csv_file, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                clean_row = {k: (v.strip() if v and v.strip() else "N/A") for k, v in row.items()}
                row_tuple = tuple(clean_row.values())
                if row_tuple not in seen:
                    seen.add(row_tuple)
                    rows.append(clean_row)

        # Keep true site order (seq)
        rows = [self.items[k] for k in sorted(self.items.keys())]

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotes"

        # Header style
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        # Borders
        border_style = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium")
        )

        # Write header
        for col_num, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col_num, value=key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border_style

        # Write data rows
        light_grey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        na_font = Font(color="808080", italic=True)

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, key in enumerate(keys, start=1):
                val = row[key] if row[key].strip() else "N/A"
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border_style
                if r_idx % 2 == 0:
                    cell.fill = light_grey
                if val == "N/A":
                    cell.font = na_font

        # ---------- Auto-fit column widths (improved) ----------
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or "")) * 1.2
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(12, min(max_length, 80))

        # Freeze header + enable filter
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # Add metadata note
        meta_row = len(rows) + 3
        note = f"📊 Sourced from (http://quotes.toscrape.com/) — {datetime.now():%Y-%m-%d %H:%M:%S}"
        meta_cell = ws.cell(row=meta_row, column=1, value=note)
        meta_cell.font = Font(color="808080", italic=True)
        ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=len(keys))

        # Save Excel
        wb.save(self.xlsx_file)
        self.logger.info(f"Wrote {len(rows)} cleaned rows to {os.path.abspath(self.xlsx_file)} ({reason})")


if __name__ == "__main__":
    process = CrawlerProcess()
    process.crawl(QuotesFinalSpider)
    process.start()
